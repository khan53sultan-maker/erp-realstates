import os
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.units import cm
from django.db.models import Sum, F
from django.utils import timezone
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_commission_report_pdf(sales, project=None, start_date=None, end_date=None):
    """
    Generate the detailed commission report as seen in the user screenshot.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#800000'), # Maroon
        alignment=1,
        spaceAfter=10
    )
    
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontSize=10,
        fontWeight='bold'
    )

    # Title
    main_title = project.name.upper() if project else "ALL PROJECTS FINANCIAL SUMMARY"
    story.append(Paragraph(main_title, title_style))
    
    sub_title = "DETAILED COMMISSION & SALES REPORT"
    story.append(Paragraph(sub_title, ParagraphStyle('SubTitle', parent=styles['Normal'], alignment=1, fontSize=12, spaceAfter=10)))
    
    period_str = "All Time"
    if start_date and end_date:
        period_str = f"{start_date} to {end_date}"
    story.append(Paragraph(f"Period: {period_str}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Calculate Top Summary Boxes (Like in screenshot)
    total_sales_val = sales.aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
    total_received_val = sum([s.total_received for s in sales])
    total_remaining_val = sum([s.current_balance for s in sales])
    
    # Extract numeric Marla from plot_size string (e.g., "5 Marla" -> 5)
    def get_marla(size_str):
        try:
            return float(''.join(filter(lambda x: x.isdigit() or x == '.', size_str)))
        except:
            return 0.0
            
    total_marla = sum([get_marla(s.plot.plot_size) for s in sales]) or 0.0

    total_landowner_comm = sales.aggregate(total=Sum('landowner_commission'))['total'] or Decimal('0.00')
    total_dealer_comm = sales.aggregate(total=Sum('dealer_commission'))['total'] or Decimal('0.00')
    total_dealer_paid = sales.aggregate(total=Sum('dealer_paid_amount'))['total'] or Decimal('0.00')
    total_dealer_remaining = total_dealer_comm - total_dealer_paid

    # Summary Table at the Top
    summary_data = [
        ['TOTAL SALE IN MARLA', f"{total_marla:.1f}", 'Total commission %', f"{total_landowner_comm:,.0f}", 'Sale partner commission', f"{total_dealer_comm:,.0f}"],
        ['PLOTS TOTAL VALUES', f"{total_sales_val:,.0f}", 'Commission Total Received', f"{total_landowner_comm:,.0f}", 'Current commission', f"{total_dealer_remaining:,.0f}"],
        ['TOTAL RECEIVED AMOUNT', f"{total_received_val:,.0f}", 'Commission Total Remaining', "0", 'Received', f"{total_dealer_paid:,.0f}"],
        ['TOTAL REMAINING AMOUNT', f"{total_remaining_val:,.0f}", '', '', 'Remaining', f"{total_dealer_remaining:,.0f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[5*cm, 3*cm, 5*cm, 3*cm, 5*cm, 3*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,3), colors.HexColor('#FFD700')), # Gold for labels
        ('BACKGROUND', (2,0), (2,2), colors.HexColor('#87CEEB')), # SkyBlue for labels
        ('BACKGROUND', (4,0), (4,3), colors.HexColor('#90EE90')), # LightGreen for labels
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 1*cm))

    # Main Data Table Headers
    # Grouping headers like in screenshot
    header_1 = [
        'S/No', 'Booking Date', 'Month', "Client's Name", 'Plot No', 'Marla', 
        'NET VALUE', 'Total Received', 'TOTAL REMAINING', 
        'Comm Received', 'SOLD BY', 'Dealer Comm', 'Paid', 'Landowner Share', 'L/O Recovered', 'L/O Paid', 'L/O Rem.', 'Status'
    ]
    
    table_data = [header_1]
    
    for i, sale in enumerate(sales, 1):
        table_data.append([
            str(i),
            sale.sale_date.strftime('%d/%m/%y') if sale.sale_date else '-',
            sale.sale_date.strftime('%b') if sale.sale_date else '-',
            sale.customer.name,
            sale.plot.plot_number,
            sale.plot.plot_size,
            f"{sale.total_price:,.0f}",
            f"{sale.total_received:,.0f}",
            f"{sale.current_balance:,.0f}",
            f"{sale.landowner_commission:,.0f}",
            sale.dealer.name if sale.dealer else '-',
            f"{sale.dealer_commission:,.0f}",
            f"{sale.dealer_paid_amount:,.0f}",
            f"{sale.landowner_total_share:,.0f}",
            f"{sale.landowner_share_received:,.0f}",
            f"{sale.landowner_paid_amount:,.0f}",
            f"{sale.landowner_share_remaining:,.0f}",
            sale.commission_status
        ])

    # Table styling
    main_table = Table(table_data, repeatRows=1)
    # Adjust col widths to fit A4 landscape (approx 27cm usable)
    # 18 columns
    col_widths = [0.7*cm, 2.0*cm, 1.1*cm, 3.2*cm, 1.3*cm, 1.3*cm, 2*cm, 2*cm, 2*cm, 1.7*cm, 1.8*cm, 1.7*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.5*cm, 1.5*cm, 1.3*cm]
    main_table._argW = col_widths

    main_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F4F4F')), # DarkSlateGray
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(main_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_sales_report_pdf(sales, start_date=None, end_date=None):
    """Simple sales report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    story.append(Paragraph("Real Estate Sales Report", styles['Heading1']))
    story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    data = [['Date', 'Project', 'Plot', 'Customer', 'Dealer', 'Price', 'Received']]
    for s in sales:
        data.append([
            s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '-',
            getattr(s.plot.project, 'name', 'N/A') if (s.plot and s.plot.project) else 'N/A',
            getattr(s.plot, 'plot_number', 'N/A') if s.plot else 'N/A',
            getattr(s.customer, 'name', 'N/A') if s.customer else 'N/A',
            getattr(s.dealer, 'name', '-') if s.dealer else '-',
            f"{float(s.total_price):,.0f}" if s.total_price is not None else "0",
            f"{float(s.total_received):,.0f}" if s.total_received is not None else "0"
        ])
    
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
    ]))
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_dealer_commission_report_pdf(sales, dealer=None, start_date=None, end_date=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    story = []
    styles = getSampleStyleSheet()
    
    title = "DEALER COMMISSION LEDGER"
    if dealer: title += f" - {dealer.name if hasattr(dealer, 'name') else dealer}"
    
    header_style = ParagraphStyle('ReportHeader', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2F4F4F'), alignment=1, spaceAfter=10)
    story.append(Paragraph(title, header_style))
    story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", styles['Italic']))
    if start_date or end_date:
        story.append(Paragraph(f"Period: {start_date or 'Start'} to {end_date or 'End'}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    data = [['Date', 'Project', 'Plot No.', 'Customer', 'Plot Price', 'Comm. Value', 'Paid so far', 'Remaining', 'Status']]
    
    total_comm = 0
    total_paid = 0
    
    for s in sales:
        rem = s.dealer_commission - s.dealer_paid_amount
        total_comm += s.dealer_commission
        total_paid += s.dealer_paid_amount
        
        data.append([
            s.sale_date.strftime('%d/%m/%y') if s.sale_date else '-',
            (s.plot.project.name[:20] if (s.plot and s.plot.project) else 'N/A'),
            (s.plot.plot_number if s.plot else 'N/A'),
            (s.customer.name[:20] if s.customer else 'N/A'),
            f"{float(s.total_price):,.0f}" if s.total_price is not None else "0",
            f"{float(s.dealer_commission):,.0f}" if s.dealer_commission is not None else "0",
            f"{float(s.dealer_paid_amount):,.0f}" if s.dealer_paid_amount is not None else "0",
            f"{float(rem):,.0f}" if rem is not None else "0",
            s.commission_status
        ])
    
    # Summary Row
    data.append(['', '', '', 'TOTALS', '', f"{total_comm:,.0f}", f"{total_paid:,.0f}", f"{(total_comm - total_paid):,.0f}", ''])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F4F4F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (3, -1), (3, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_landowner_payout_report_pdf(sales, project=None, start_date=None, end_date=None):
    """
    Specific ledger for payments made to the Landowner (Malik).
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    story = []
    styles = getSampleStyleSheet()
    
    title = "LANDOWNER (MALIK) SETTLEMENT LEDGER"
    if project: title += f" - {project.name}"
    
    header_style = ParagraphStyle('ReportHeader', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4B0082'), alignment=1, spaceAfter=10) # Indigo
    story.append(Paragraph(title, header_style))
    story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", styles['Italic']))
    story.append(Spacer(1, 0.5*cm))
    
    data = [['Date', 'Project', 'Plot No.', 'Customer', 'Net Plot Value', 'L/O Total Share', 'L/O Recovered', 'L/O Paid (Cash)', 'L/O Balance']]
    
    totals = {
        'share': Decimal('0'),
        'recovered': Decimal('0'),
        'paid': Decimal('0'),
        'remaining': Decimal('0')
    }
    
    for s in sales:
        totals['share'] += Decimal(str(s.landowner_total_share))
        totals['recovered'] += Decimal(str(s.landowner_share_received))
        totals['paid'] += Decimal(str(s.landowner_paid_amount))
        totals['remaining'] += Decimal(str(s.landowner_share_remaining))
        
        data.append([
            s.sale_date.strftime('%d/%m/%y') if s.sale_date else '-',
            (s.plot.project.name[:20] if (s.plot and s.plot.project) else 'N/A'),
            (s.plot.plot_number if s.plot else 'N/A'),
            (s.customer.name[:20] if s.customer else 'N/A'),
            f"{float(s.total_price):,.0f}" if s.total_price is not None else "0",
            f"{float(s.landowner_total_share):,.0f}" if s.landowner_total_share is not None else "0",
            f"{float(s.landowner_share_received):,.0f}" if s.landowner_share_received is not None else "0",
            f"{float(s.landowner_paid_amount):,.0f}" if s.landowner_paid_amount is not None else "0",
            f"{float(s.landowner_share_remaining):,.0f}" if s.landowner_share_remaining is not None else "0"
        ])
    
    # Summary Row
    data.append(['', '', '', 'TOTALS', '', f"{totals['share']:,.0f}", f"{totals['recovered']:,.0f}", f"{totals['paid']:,.0f}", f"{totals['remaining']:,.0f}"])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4B0082')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_client_payment_report_pdf(sales, start_date=None, end_date=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    story = []
    styles = getSampleStyleSheet()
    
    story.append(Paragraph("Client Payment & Due Report", styles['Heading1']))
    story.append(Spacer(1, 0.5*cm))
    
    data = [['Date', 'Client', 'Phone', 'Project', 'Plot No.', 'Total Price', 'Received', 'Balance', 'Status']]
    for s in sales:
        data.append([
            s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '-',
            s.customer.name,
            s.customer.phone,
            s.plot.project.name,
            s.plot.plot_number,
            f"{s.total_price:,.0f}",
            f"{s.total_received:,.0f}",
            f"{s.current_balance:,.0f}",
            s.commission_status
        ])
        
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_customer_ledger_pdf(ledger_entries, summary, start_date=None, end_date=None):
    """Premium Customer Ledger PDF matching the UI"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    story = []
    styles = getSampleStyleSheet()
    
    # Header
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, textColor=colors.HexColor('#7D0541'))
    story.append(Paragraph("CUSTOMER TRANSACTION LEDGER", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Customer Info Table
    info_data = [
        [Paragraph(f"<b>Customer:</b> {summary['customer_name']}", styles['Normal']), 
         Paragraph(f"<b>Phone:</b> {summary['customer_phone']}", styles['Normal'])],
        [Paragraph(f"<b>Period:</b> {start_date or 'Start'} to {end_date or 'End'}", styles['Normal']),
         Paragraph(f"<b>Outstanding Balance:</b> Rs.{summary['outstanding_balance']:,.2f}", styles['Normal'])]
    ]
    info_table = Table(info_data, colWidths=[9*cm, 9*cm])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Ledger Table
    data = [['Date', 'Description', 'Type', 'Debit', 'Credit', 'Balance']]
    
    for e in ledger_entries:
        data.append([
            e['date'],
            Paragraph(e['description'], styles['Normal']),
            e['type'],
            f"{e['debit']:,.2f}" if e['debit'] > 0 else "-",
            f"{e['credit']:,.2f}" if e['credit'] > 0 else "-",
            f"{e['balance']:,.2f}"
        ])
    
    # Table Styles
    t = Table(data, colWidths=[2.2*cm, 7.5*cm, 2.5*cm, 2.3*cm, 2.3*cm, 2.3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7D0541')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'), # Amounts to right
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white])
    ]))
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_payment_plan_pdf(sale, installments):
    """Premium Payment Plan PDF matching Excel"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    story = []
    styles = getSampleStyleSheet()
    
    # Header
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, textColor=colors.HexColor('#2F4F4F'))
    story.append(Paragraph(f"PAYMENT PLAN - {sale.customer.name.upper()}", title_style))
    story.append(Spacer(1, 0.2*cm))
    
    proj_name = getattr(sale.plot.project, 'name', 'N/A') if (sale.plot and sale.plot.project) else 'N/A'
    plot_num = getattr(sale.plot, 'plot_number', 'N/A') if sale.plot else 'N/A'
    total_px = float(sale.total_price) if sale.total_price is not None else 0.0
    
    info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], alignment=1, fontWeight='bold')
    story.append(Paragraph(f"Project: {proj_name} | Plot: {plot_num} | Price: Rs.{total_px:,.0f}", info_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Data Table
    data = [['Month', 'Due Amount', 'Receipt Date', 'Receipt No', 'Paid Amount', 'Remaining']]
    
    # Add Down Payment
    dp_val = sale.down_payment or 0
    dp_recd = sale.received_down_payment or 0
    dp_rem = dp_val - dp_recd
    data.append([
        "Down Payment",
        f"{dp_val:,.0f}",
        sale.sale_date.strftime('%d/%m/%Y') if sale.sale_date else '-',
        "DP-BOOKING",
        f"{dp_recd:,.0f}",
        f"{dp_rem:,.0f}"
    ])
    
    total_paid = dp_recd
    for inst in installments:
        inst_amt = inst.amount or 0
        inst_paid = inst.paid_amount or 0
        rem = inst_amt - inst_paid
        month_label = inst.due_date.strftime('%B %Y') if inst.due_date else 'N/A'
        data.append([
            month_label,
            f"{inst_amt:,.0f}",
            inst.paid_date.strftime('%d/%m/%Y') if inst.paid_date else '-',
            f"REC-{str(inst.id)[:5].upper()}" if inst.paid_date else '-',
            f"{inst_paid:,.0f}",
            f"{rem:,.0f}"
        ])
        total_paid += inst_paid
    
    t = Table(data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F4F4F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#87CEEB')), # DP Row
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    story.append(t)
    
    # Totals
    story.append(Spacer(1, 0.5*cm))
    summary_data = [
        ['TOTAL PRICE', f"Rs.{sale.total_price:,.0f}"],
        ['TOTAL RECEIVED', f"Rs.{total_paid:,.0f}"],
        ['REMAINING BALANCE', f"Rs.{sale.total_price - total_paid:,.0f}"]
    ]
    st = Table(summary_data, colWidths=[4*cm, 3*cm])
    st.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
    ]))
    story.append(st)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_profit_loss_report_pdf(incomes, expenses, start_date=None, end_date=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    header_style = ParagraphStyle('ReportHeader', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1A5276'), alignment=1)
    story.append(Paragraph("PROFIT AND LOSS STATEMENT", header_style))
    story.append(Spacer(1, 0.5*cm))
    
    if start_date and end_date:
        story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Income Section
    story.append(Paragraph("INCOME", styles['Heading2']))
    income_data = [['Description', 'Amount']]
    income_groups = {}
    for inc in incomes:
        etype = inc.get_income_type_display()
        income_groups[etype] = income_groups.get(etype, 0) + inc.amount
    
    for etype, amount in income_groups.items():
        income_data.append([etype, f"{amount:,.0f}"])
    
    total_inc = sum(income_groups.values())
    income_data.append([Paragraph('<b>TOTAL INCOME</b>', styles['Normal']), f"<b>{total_inc:,.0f}</b>"])
    
    it = Table(income_data, colWidths=[12*cm, 4*cm])
    it.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4E6F1')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(it)
    story.append(Spacer(1, 0.5*cm))

    # Expense Section
    story.append(Paragraph("EXPENSES", styles['Heading2']))
    expense_data = [['Description', 'Amount']]
    expense_groups = {}
    for exp in expenses:
        ecat = exp.get_category_display()
        expense_groups[ecat] = expense_groups.get(ecat, 0) + exp.amount
        
    for ecat, amount in expense_groups.items():
        expense_data.append([ecat, f"{amount:,.0f}"])
        
    total_exp = sum(expense_groups.values())
    expense_data.append([Paragraph('<b>TOTAL EXPENSES</b>', styles['Normal']), f"<b>{total_exp:,.0f}</b>"])
    
    et = Table(expense_data, colWidths=[12*cm, 4*cm])
    et.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FADBD8')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(et)
    story.append(Spacer(1, 1*cm))

    # Net Profit
    profit = total_inc - total_exp
    profit_color = colors.green if profit >= 0 else colors.red
    data = [['NET PROFIT / LOSS', f"{profit:,.0f}"]]
    pt = Table(data, colWidths=[12*cm, 4*cm])
    pt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (1, 0), (1, 0), profit_color),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    story.append(pt)

    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_cash_flow_report_pdf(transactions, start_date=None, end_date=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    story = []
    styles = getSampleStyleSheet()
    
    story.append(Paragraph("Cash Flow Report", styles['Heading1']))
    story.append(Spacer(1, 0.5*cm))
    
    data = [['Date', 'Type', 'Category', 'Description', 'Income Amount', 'Expense Amount']]
    net = 0
    total_in = 0
    total_out = 0
    for t in transactions:
        amount = t['amount']
        if t['type'] == 'Income':
            total_in += amount
            net += amount
            data.append([t['date'].strftime('%Y-%m-%d') if t['date'] else '-', 'Income', t['category'], t.get('description', ''), f"{amount:,.0f}", "-"])
        else:
            total_out += amount
            net -= amount
            data.append([t['date'].strftime('%Y-%m-%d') if t['date'] else '-', 'Expense', t['category'], t.get('description', ''), "-", f"{amount:,.0f}"])
            
    data.append(["TOTAL", "", "", "", f"{total_in:,.0f}", f"{total_out:,.0f}"])
    data.append(["NET CASH FLOW", "", "", "", f"{net:,.0f}", ""])
            
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_commission_report_excel(sales, project=None, start_date=None, end_date=None):
    """
    Generate the High-End Stylized Commission Report with Summary Blocks and Merged Headers.
    Matches the user's latest screenshot.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from decimal import Decimal
    from django.db.models import Sum
    from io import BytesIO
    import re

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    def write_sheet(ws, project_sales, project_obj):
        # Styles
        GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        SKY_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        LIGHT_GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        DARK_FILL = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        
        BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        BOLD_FONT = Font(bold=True, size=11)
        WHITE_FONT = Font(color="FFFFFF", bold=True)
        
        # Calculate Marla Sum
        total_marla = 0.0
        for s in project_sales:
            size_str = s.plot.plot_size if s.plot and s.plot.plot_size else "0"
            # Extract numbers from string like "5 Marla"
            nums = re.findall(r'\d+\.?\d*', size_str)
            if nums: total_marla += float(nums[0])

        # 1. Top Summary Blocks (Rows 1-4)
        total_sales_val = project_sales.aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
        total_received_val = sum([s.total_received for s in project_sales]) or Decimal('0.00')
        total_remaining_val = sum([s.current_balance for s in project_sales]) or Decimal('0.00')
        
        total_land_comm = project_sales.aggregate(total=Sum('landowner_commission'))['total'] or Decimal('0.00')
        total_land_recv = project_sales.aggregate(total=Sum('landowner_commission_received'))['total'] or Decimal('0.00')
        
        total_dealer_comm = project_sales.aggregate(total=Sum('dealer_commission'))['total'] or Decimal('0.00')
        total_dealer_earned = sum([s.current_dealer_commission for s in project_sales]) or Decimal('0.00')
        total_dealer_paid = project_sales.aggregate(total=Sum('dealer_paid_amount'))['total'] or Decimal('0.00')

        # Block 1: Sales (Gold)
        summary_left = [
            ('TOTAL SALE IN MARLA', total_marla),
            ('PLOTS TOTAL VALUES', total_sales_val),
            ('TOTAL RECEIVED AMOUNT', total_received_val),
            ('TOTAL REMAINING AMOUNT', total_remaining_val),
        ]
        for r_idx, (lbl, val) in enumerate(summary_left, 1):
            ws.cell(row=r_idx, column=1, value=lbl).fill = GOLD_FILL
            ws.cell(row=r_idx, column=1).font = BOLD_FONT
            ws.cell(row=r_idx, column=1).border = BORDER
            vc = ws.cell(row=r_idx, column=2, value=val)
            vc.border = BORDER; vc.number_format = '#,##0'

        # Block 2: Commission (Blue)
        summary_mid = [
            ('Total commission %', total_land_comm),
            ('Commission Total Received', total_land_recv),
            ('Commission Total Remaining', total_land_comm - total_land_recv),
        ]
        for r_idx, (lbl, val) in enumerate(summary_mid, 1):
            c_label = ws.cell(row=r_idx, column=11, value=lbl)
            c_label.fill = SKY_BLUE_FILL; c_label.font = BOLD_FONT; c_label.border = BORDER
            c_val = ws.cell(row=r_idx, column=12, value=val)
            c_val.border = BORDER; c_val.number_format = '#,##0'

        # Block 3: Sale Partner (Green) - Fixed Column Q=17, R=18
        summary_right = [
            ('Sale partner commission', total_dealer_comm),
            ('Current commission', total_dealer_earned),
            ('Received', total_dealer_paid),
            ('Remaining', total_dealer_comm - total_dealer_paid),
        ]
        for r_idx, (lbl, val) in enumerate(summary_right, 1):
            c_label = ws.cell(row=r_idx, column=17, value=lbl)
            c_label.fill = LIGHT_GREEN_FILL; c_label.font = BOLD_FONT; c_label.border = BORDER
            c_val = ws.cell(row=r_idx, column=18, value=val)
            c_val.border = BORDER; c_val.number_format = '#,##0'

        # 2. Category Row (Row 6)
        ws.merge_cells('H6:J6'); ws['H6'] = 'CLIENT AND PLOTS DETAIL'; ws['H6'].fill = YELLOW_FILL
        ws.merge_cells('K6:M6'); ws['K6'] = 'Commission Details'; ws['K6'].fill = YELLOW_FILL
        ws.merge_cells('Q6:T6'); ws['Q6'] = 'Sale partner commission'; ws['Q6'].fill = YELLOW_FILL
        for c in [8, 11, 17]:
            ws.cell(row=6, column=c).alignment = Alignment(horizontal='center'); ws.cell(row=6, column=c).font = BOLD_FONT
            ws.cell(row=6, column=c).border = BORDER

        # 3. Column Headers (Row 7)
        headers = [
            'S/No', 'Booking Date', 'Month', "Client's Name", 'Reg No', 'Plot No', 'Marla',
            'NET VALUE', 'Total Received', 'TOTAL REMAINING',
            'Total commission %', 'Total Received', 'Total Remaining',
            'SOLD BY', 'Downpayment', 'Received Down',
            'Sale partner commission', 'Current commission', 'Received', 'Remaining', 'Status'
        ]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=c_idx, value=h)
            cell.font = WHITE_FONT; cell.fill = DARK_FILL; cell.border = BORDER; cell.alignment = Alignment(horizontal='center')

        # 4. Data Rows (Row 8 onwards)
        for r_idx, sale in enumerate(project_sales, 8):
            row_data = [
                r_idx - 7,
                sale.sale_date.strftime('%d/%m/%y') if sale.sale_date else '',
                sale.sale_date.strftime('%b') if sale.sale_date else '',
                sale.customer.name.upper() if sale.customer else '',
                sale.registration_number or '',
                sale.plot.plot_number if sale.plot else '',
                sale.plot.plot_size if sale.plot else '',
                sale.total_price,
                sale.total_received,
                sale.current_balance,
                sale.landowner_commission,
                sale.landowner_commission_received,
                sale.landowner_commission_remaining,
                sale.dealer.name if sale.dealer else '-',
                sale.down_payment,
                sale.received_down_payment,
                sale.dealer_commission,
                sale.current_dealer_commission,
                sale.dealer_paid_amount,
                sale.dealer_commission_remaining,
                sale.commission_status
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = BORDER
                if isinstance(val, (int, float, Decimal)):
                    cell.number_format = '#,##0'
                # Coloring data sections
                if 8 <= c_idx <= 10: cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Gold Tint
                if 11 <= c_idx <= 13: cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Blue Tint
                if 17 <= c_idx <= 20: cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Green Tint

        # Auto-width
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

    # Main Sheet Logic
    if project:
        ws = wb.create_sheet(title=project.name[:30])
        write_sheet(ws, sales, project)
    else:
        from .models import Project
        p_ids = sales.values_list('plot__project_id', flat=True).distinct()
        all_projects = Project.objects.filter(id__in=p_ids)
        if not all_projects.exists():
            ws = wb.create_sheet(title="Report")
            write_sheet(ws, sales, None)
        else:
            for p in all_projects:
                ws = wb.create_sheet(title=p.name[:30])
                p_sales = sales.filter(plot__project=p)
                write_sheet(ws, p_sales, p)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_payment_plan_excel(sale, installments):
    """
    Generate the detailed Payment Plan spreadsheet for a specific sale.
    Matches the user's screenshot layout.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from decimal import Decimal
    from io import BytesIO
    from django.utils import timezone
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Plan"

    # Define Styles
    GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    SKY_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    DARK_FILL = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    RED_FONT = Font(color="FF0000", bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header Row (Sale Info)
    cust_name = getattr(sale.customer, 'name', 'N/A').upper()
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value=f"PAYMENT PLAN - {cust_name}")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    proj_name = getattr(sale.plot.project, 'name', 'N/A') if (sale.plot and sale.plot.project) else 'N/A'
    plot_num = getattr(sale.plot, 'plot_number', 'N/A') if sale.plot else 'N/A'
    total_px = float(sale.total_price) if sale.total_price is not None else 0.0
    
    ws.merge_cells('A2:G2')
    info_cell = ws.cell(row=2, column=1, value=f"Project: {proj_name} | Plot: {plot_num} | Price: Rs.{total_px:,.0f}")
    info_cell.font = BOLD_FONT
    info_cell.alignment = Alignment(horizontal='center')

    # Column Headers (Row 4)
    headers = ['Month', 'Due Amount', 'Receipt Date', 'Receipt No', 'Paid Amount', 'Remaining', 'Overdue']
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c_idx, value=header)
        cell.fill = DARK_FILL
        cell.font = WHITE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    # Data Rows
    current_row = 5
    total_due = Decimal('0')
    total_paid = Decimal('0')
    total_overdue = Decimal('0')

    # 1. Add Down Payment Row
    dp_rem = sale.down_payment - sale.received_down_payment
    dp_data = [
        "Down Payment",
        sale.down_payment,
        sale.sale_date.strftime('%d/%m/%Y') if sale.sale_date else '-',
        "DP-BOOKING",
        sale.received_down_payment,
        dp_rem,
        Decimal('0')
    ]
    for c_idx, val in enumerate(dp_data, 1):
        cell = ws.cell(row=current_row, column=c_idx, value=val)
        cell.fill = SKY_BLUE_FILL
        cell.border = BORDER
        if isinstance(val, (int, float, Decimal)): cell.number_format = '#,##0'

    total_due += sale.down_payment
    total_paid += sale.received_down_payment
    current_row += 1

    # 2. Add Installment Rows
    for inst in installments:
        rem = inst.amount - inst.paid_amount
        is_overdue = inst.status != 'PAID' and inst.due_date < timezone.now().date()
        overdue_amt = rem if is_overdue else Decimal('0')

        # Distinguish Balloon rows
        base_inst = sale.installment_amount if sale.installment_amount else Decimal('1.0')
        is_balloon = (inst.payment_remarks and 'BALLOON' in inst.payment_remarks.upper()) or (inst.amount > base_inst * Decimal('1.2'))
        
        month_label = inst.due_date.strftime('%B %Y')
        if is_balloon:
            month_label += " (Balloon Payment)"

        data = [
            month_label,
            inst.amount,
            inst.paid_date.strftime('%d/%m/%Y') if inst.paid_date else '-',
            f"REC-{str(inst.id)[:5].upper()}" if inst.paid_date else '-',
            inst.paid_amount,
            rem,
            overdue_amt
        ]

        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float, Decimal)):
                cell.number_format = '#,##0'
                if c_idx == 7 and val > 0: cell.font = RED_FONT
            
        if is_balloon:
            for c_idx in range(1, 8):
                ws.cell(row=current_row, column=c_idx).fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                ws.cell(row=current_row, column=c_idx).font = Font(bold=True)

        total_due += inst.amount
        total_paid += inst.paid_amount
        total_overdue += overdue_amt
        current_row += 1

    # Totals Section
    current_row += 1
    summary_start = current_row
    
    ws.cell(row=current_row, column=5, value="TOTAL PRICE").font = BOLD_FONT
    ws.cell(row=current_row, column=6, value=sale.total_price).font = BOLD_FONT
    ws.cell(row=current_row, column=6).number_format = '#,##0'
    current_row += 1

    ws.cell(row=current_row, column=5, value="TOTAL RECEIVED").font = BOLD_FONT
    ws.cell(row=current_row, column=6, value=total_paid).font = BOLD_FONT
    ws.cell(row=current_row, column=6).number_format = '#,##0'
    current_row += 1

    ws.cell(row=current_row, column=5, value="REMAINING BALANCE").font = BOLD_FONT
    ws.cell(row=current_row, column=6, value=sale.total_price - total_paid).font = BOLD_FONT
    ws.cell(row=current_row, column=6).number_format = '#,##0'
    current_row += 1

    ws.cell(row=current_row, column=5, value="TOTAL OVERDUE").font = BOLD_FONT
    ws.cell(row=current_row, column=6, value=total_overdue).font = RED_FONT
    ws.cell(row=current_row, column=6).number_format = '#,##0'

    # Stylize summary box
    for r in range(summary_start, current_row + 1):
        ws.cell(row=r, column=5).border = BORDER
        ws.cell(row=r, column=6).border = BORDER

    # Auto-fit
    for col_idx in range(1, 8):
        max_length = 0
        column = get_column_letter(col_idx)
        for row_idx in range(1, current_row + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_executive_statement_excel(sale, installments):
    """
    Final Executive Statement fix - 100% matched to Step 553 screenshot.
    - Uses Project Name in Header.
    - Defensive formatting for None types.
    - RE-ENABLED Gridlines for standard Excel look.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from decimal import Decimal
    from io import BytesIO
    from datetime import datetime
    import logging

    logger = logging.getLogger(__name__)

    def fmt_str(val, title=False):
        if val is None: return ""
        s = str(val).strip()
        return s.title() if title else s

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Statement"
        # Re-enabling gridlines as per user request
        ws.sheet_view.showGridLines = True

        # Define Styles
        BOLD_FONT = Font(bold=True, size=11)
        TITLE_FONT = Font(bold=True, size=14)
        CENTER = Alignment(horizontal='center', vertical='center')
        LEFT = Alignment(horizontal='left', vertical='center')
        RIGHT = Alignment(horizontal='right', vertical='center')
        THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        MEDIUM = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        
        # Precise Column Widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        # Row 1: PROJECT NAME
        project_name = "PROJECT NAME"
        if getattr(sale, 'plot', None) and getattr(sale.plot, 'project', None):
            project_name = sale.plot.project.name.upper()
        
        ws.merge_cells('A1:B1')
        ws['A1'] = project_name
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER
        ws['A1'].border = MEDIUM

        # Rows 2-13: Personal Details
        cust = getattr(sale, 'customer', None)
        
        details = [
            ('Reg No#', fmt_str(getattr(sale, 'registration_number', ''))),
            ('Date', sale.sale_date.strftime('%d-%b-%y') if getattr(sale, 'sale_date', None) else ''),
            ('Name', fmt_str(getattr(cust, 'name', ''), title=True)),
            ('F/Name', fmt_str(getattr(cust, 'father_name', ''), title=True)),
            ('Contact No#', fmt_str(getattr(cust, 'phone', ''))),
            ('CNIC No#', fmt_str(getattr(cust, 'cnic', ''))),
            ('Address', fmt_str(getattr(cust, 'address', ''))),
            ('Block Name', fmt_str(getattr(sale, 'block_name', ''))),
            ('Plot No#', fmt_str(sale.plot.plot_number if getattr(sale, 'plot', None) else '')),
            ('Plot Size', fmt_str(sale.plot.plot_size if getattr(sale, 'plot', None) else '')),
        ]
        
        for i, (lbl, val) in enumerate(details, 2):
            cell_lbl = ws.cell(row=i, column=1, value=lbl)
            cell_lbl.font = BOLD_FONT
            cv = ws.cell(row=i, column=2, value=val)
            cv.alignment = LEFT
            if lbl in ['Contact No#', 'CNIC No#', 'Address']:
                cv.border = THIN

        # Row 14: Installment PLAN Header
        ws.merge_cells('A14:B14')
        ws['A14'] = "Installment PLAN"
        ws['A14'].font = BOLD_FONT
        ws['A14'].alignment = CENTER

        # Rows 15-22: Plan Items
        alc = float(getattr(sale, 'allocation_amount', 0) or 0)
        conf = float(getattr(sale, 'confirmation_amount', 0) or 0)
        dp = float(getattr(sale, 'down_payment', 0) or 0)
        tp = float(getattr(sale, 'total_price', 0) or 0)
        poss = float(getattr(sale, 'possession_amount', 0) or 0)
        proc = float(getattr(sale, 'processing_amount', 0) or 0)
        lp = float(getattr(sale, 'last_payment_amount', 0) or 0)
        
        booking_val = dp - (alc + conf)
        inst_total = tp - (dp + poss + proc + lp)
        
        plan_rows = [
            ('Booking', booking_val),
            ('Allocation', alc),
            ('Confirmation', conf),
            ('Installments', inst_total),
            ('Processing Fee', proc),
            ('Possession', poss),
            ('Last Payment', lp)
        ]
        
        for i, (lbl, amt) in enumerate(plan_rows, 15):
            ws.cell(row=i, column=1, value=lbl).font = BOLD_FONT
            val_cell = ws.cell(row=i, column=3, value=amt)
            val_cell.number_format = '#,##0.00'; val_cell.alignment = RIGHT
            if lbl in ['Possession', 'Add Extra']:
                ws.cell(row=i, column=2, value="-").alignment = CENTER

        # Row 23: Plan Total
        ws.cell(row=23, column=1, value="Total").font = BOLD_FONT
        ws.cell(row=23, column=1).border = MEDIUM
        t_val = ws.cell(row=23, column=2, value=tp)
        t_val.font = BOLD_FONT; t_val.number_format = '#,##0.00'; t_val.alignment = RIGHT; t_val.border = MEDIUM

        # Row 25: Ledger Header
        ws.merge_cells('A25:D25')
        ws['A25'] = "Receipt's / Payment Detail"; ws['A25'].font = BOLD_FONT; ws['A25'].alignment = CENTER

        # Row 26: Ledger Sub-Headers
        columns = ['Receipt No#', 'Installment Date', 'Amount', 'Rem/Balance']
        for c_idx, h in enumerate(columns, 1):
            cell = ws.cell(row=26, column=c_idx, value=h); cell.font = BOLD_FONT; cell.border = THIN; cell.alignment = CENTER

        from .models import InstallmentPayment
        all_payments = list(InstallmentPayment.objects.filter(installment__sale=sale).order_by('payment_date', 'created_at'))
        
        row_cursor = 27
        total_paid = Decimal('0.00'); rem_bal = Decimal(str(tp))
        
        # 1. Downpayment
        rcvd_dp = Decimal(str(getattr(sale, 'received_down_payment', 0) or 0))
        if rcvd_dp > 0:
            total_paid += rcvd_dp; rem_bal -= rcvd_dp
            receipt_val = sale.receipt_number if sale.receipt_number else f"BK-{str(sale.id)[:5].upper()}"
            ws.cell(row=row_cursor, column=1, value=receipt_val).border = THIN; ws.cell(row=row_cursor, column=1).alignment = CENTER
            ws.cell(row=row_cursor, column=2, value=sale.sale_date.strftime('%d-%b-%y') if getattr(sale, 'sale_date', None) else '').border = THIN; ws.cell(row=row_cursor, column=2).alignment = CENTER
            c3 = ws.cell(row=row_cursor, column=3, value=float(rcvd_dp)); c3.number_format = '#,##0.00'; c3.border = THIN; c3.alignment = RIGHT
            c4 = ws.cell(row=row_cursor, column=4, value=float(rem_bal)); c4.number_format = '#,##0.00'; c4.border = THIN; c4.alignment = RIGHT
            row_cursor += 1

        # 2. Subsequent Payments
        receipt_counter = 2
        for p in all_payments:
            if row_cursor > 36: break
            p_amt = Decimal(str(p.amount)); total_paid += p_amt; rem_bal -= p_amt
            receipt_val = p.receipt_number if p.receipt_number else f"RE-{str(p.id)[:5].upper()}"
            ws.cell(row=row_cursor, column=1, value=receipt_val).border = THIN; ws.cell(row=row_cursor, column=1).alignment = CENTER
            ws.cell(row=row_cursor, column=2, value=p.payment_date.strftime('%d-%b-%y') if p.payment_date else '').border = THIN; ws.cell(row=row_cursor, column=2).alignment = CENTER
            c3 = ws.cell(row=row_cursor, column=3, value=float(p_amt)); c3.number_format = '#,##0.00'; c3.border = THIN; c3.alignment = RIGHT
            c4 = ws.cell(row=row_cursor, column=4, value=float(rem_bal)); c4.number_format = '#,##0.00'; c4.border = THIN; c4.alignment = RIGHT
            row_cursor += 1; receipt_counter += 1

        while row_cursor <= 36:
            ws.cell(row=row_cursor, column=1).border = THIN
            ws.cell(row=row_cursor, column=2).border = THIN
            ws.cell(row=row_cursor, column=3).border = THIN
            ws.cell(row=row_cursor, column=4).border = THIN
            row_cursor += 1

        # Row 37: Total
        ws.merge_cells('A37:B37'); ws['A37'] = "Total"; ws['A37'].font = BOLD_FONT; ws['A37'].alignment = CENTER; ws['A37'].border = MEDIUM
        ws.cell(row=37, column=1).border = MEDIUM; ws.cell(row=37, column=2).border = MEDIUM
        total_val_cell = ws.cell(row=37, column=3, value=float(total_paid)); total_val_cell.font = BOLD_FONT; total_val_cell.alignment = RIGHT; total_val_cell.number_format = '#,##0.00'; total_val_cell.border = MEDIUM
        ws.cell(row=37, column=4).border = MEDIUM

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        logger.error(f"Error generating executive statement: {str(e)}")
        raise e

def generate_cash_flow_report_excel(transactions, start_date=None, end_date=None):
    """
    Generate the Standardized Cash Flow Report in Excel format.
    Matches PDF format: Date | Type | Category | Description | Income Amount | Expense Amount
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from decimal import Decimal

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow Report"

    # Styles
    MAROON_FILL = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
    BOLD = Font(bold=True, size=11)
    CENTER = Alignment(horizontal='center', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "CASH FLOW REPORT"
    ws['A1'].font = Font(size=16, bold=True, color="7D0541")
    ws['A1'].alignment = CENTER

    if start_date and end_date:
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].alignment = CENTER
        ws['A2'].font = BOLD

    # Table Headers (Row 4)
    headers = ['Date', 'Type', 'Category', 'Description', 'Income Amount', 'Expense Amount']
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = WHITE_BOLD
        cell.fill = MAROON_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    # Data
    row_idx = 5
    total_income = Decimal('0')
    total_expense = Decimal('0')

    for t in transactions:
        amt = Decimal(str(t['amount']))
        is_income = t['type'] == 'Income'
        
        row_data = [
            t['date'].strftime('%Y-%m-%d') if hasattr(t['date'], 'strftime') else str(t['date']),
            t['type'],
            t['category'],
            t.get('description', ''),
            float(amt) if is_income else 0,
            float(amt) if not is_income else 0
        ]
        
        if is_income: total_income += amt
        else: total_expense += amt

        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.border = BORDER
            if c_idx >= 5:
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT
                if val == 0: cell.value = "-"
            else:
                cell.alignment = CENTER if c_idx <= 2 else Alignment(horizontal='left')
        
        row_idx += 1

    # Totals (Row idx after loop)
    ws.cell(row=row_idx, column=1, value="TOTALS").font = BOLD
    ws.cell(row=row_idx, column=5, value=float(total_income)).font = BOLD
    ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
    ws.cell(row=row_idx, column=6, value=float(total_expense)).font = BOLD
    ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
    
    for c in range(1, 7):
        ws.cell(row=row_idx, column=c).border = BORDER
        if c >= 5: ws.cell(row=row_idx, column=c).fill = GOLD_FILL
    row_idx += 1

    # Net Cash Flow
    ws.cell(row=row_idx, column=1, value="NET CASH FLOW").font = BOLD
    net_val = ws.cell(row=row_idx, column=5, value=float(total_income - total_expense))
    net_val.font = Font(bold=True, size=12, color="008000" if (total_income - total_expense) >= 0 else "FF0000")
    net_val.number_format = '#,##0.00'
    
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    for c in range(1, 7):
        ws.cell(row=row_idx, column=c).border = BORDER
    row_idx += 1

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_profit_loss_report_excel(incomes, expenses, start_date=None, end_date=None):
    """
    Generate the Standardized Profit and Loss Statement in Excel.
    Premium design with Maroon/Gold highlights.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from decimal import Decimal

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit and Loss"

    MAROON_FILL = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    SKY_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    PEACH_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    BOLD = Font(bold=True, size=11)
    WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)

    ws.merge_cells('A1:B1')
    ws['A1'] = "PROFIT AND LOSS STATEMENT"
    ws['A1'].font = Font(size=14, bold=True, color="7D0541")
    ws['A1'].alignment = Alignment(horizontal='center')

    curr_row = 3
    
    # Income Section
    ws.cell(row=curr_row, column=1, value="INCOME").font = BOLD
    ws.cell(row=curr_row, column=1).fill = GOLD_FILL
    curr_row += 1
    
    headers = ['Description', 'Amount']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=curr_row, column=i, value=h)
        cell.font = WHITE_BOLD; cell.fill = MAROON_FILL; cell.border = BORDER
    curr_row += 1

    income_groups = {}
    for inc in incomes:
        t = inc.get_income_type_display()
        income_groups[t] = income_groups.get(t, Decimal('0')) + inc.amount
    
    total_inc = Decimal('0')
    for t, amt in income_groups.items():
        ws.cell(row=curr_row, column=1, value=t).border = BORDER
        c2 = ws.cell(row=curr_row, column=2, value=float(amt))
        c2.number_format = '#,##0.00'; c2.border = BORDER; c2.alignment = Alignment(horizontal='right')
        c2.fill = SKY_BLUE_FILL
        total_inc += amt
        curr_row += 1
        
    ws.cell(row=curr_row, column=1, value="TOTAL INCOME").font = BOLD
    ws.cell(row=curr_row, column=1).border = BORDER
    c_tot = ws.cell(row=curr_row, column=2, value=float(total_inc))
    c_tot.font = BOLD; c_tot.number_format = '#,##0.00'; c_tot.border = BORDER; c_tot.fill = GOLD_FILL; c_tot.alignment = Alignment(horizontal='right')
    curr_row += 2

    # Expense Section
    ws.cell(row=curr_row, column=1, value="EXPENSES").font = BOLD
    ws.cell(row=curr_row, column=1).fill = GOLD_FILL
    curr_row += 1
    
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=curr_row, column=i, value=h)
        cell.font = WHITE_BOLD; cell.fill = MAROON_FILL; cell.border = BORDER
    curr_row += 1

    expense_groups = {}
    for exp in expenses:
        t = exp.get_category_display()
        expense_groups[t] = expense_groups.get(t, Decimal('0')) + exp.amount
        
    total_exp = Decimal('0')
    for t, amt in expense_groups.items():
        ws.cell(row=curr_row, column=1, value=t).border = BORDER
        c2 = ws.cell(row=curr_row, column=2, value=float(amt))
        c2.number_format = '#,##0.00'; c2.border = BORDER; c2.alignment = Alignment(horizontal='right')
        c2.fill = PEACH_FILL
        total_exp += amt
        curr_row += 1
        
    ws.cell(row=curr_row, column=1, value="TOTAL EXPENSES").font = BOLD
    ws.cell(row=curr_row, column=1).border = BORDER
    c_tot = ws.cell(row=curr_row, column=2, value=float(total_exp))
    c_tot.font = BOLD; c_tot.number_format = '#,##0.00'; c_tot.border = BORDER; c_tot.fill = GOLD_FILL; c_tot.alignment = Alignment(horizontal='right')
    curr_row += 2

    # Summary
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
    net = total_inc - total_exp
    ws.cell(row=curr_row, column=1, value=f"NET PROFIT / LOSS: Rs. {float(net):,.2f}").font = Font(bold=True, size=12, color="008000" if net >= 0 else "FF0000")
    ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=curr_row, column=1).border = BORDER
    ws.cell(row=curr_row, column=2).border = BORDER
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_sales_report_excel(sales, start_date=None, end_date=None):
    """Premium Sales Report Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    MAROON = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    WHITE_BOLD = Font(color="FFFFFF", bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:L1')
    ws['A1'] = "DETAILED SALES REPORT"
    ws['A1'].font = Font(size=14, bold=True, color="7D0541")
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'S/No', 'Date', 'Month', 'Client Name', 'Reg No', 'Plot No', 'Size',
        'Total Price', 'Received', 'Balance', 'Sold By', 'Status'
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = WHITE_BOLD; c.fill = MAROON; c.border = BORDER; c.alignment = Alignment(horizontal='center')

    for r_idx, s in enumerate(sales, 4):
        data = [
            r_idx - 3,
            s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '-',
            s.sale_date.strftime('%b') if s.sale_date else '',
            s.customer.name.upper() if s.customer else '-',
            s.registration_number or '-',
            s.plot.plot_number if s.plot else '-',
            s.plot.plot_size if s.plot else '-',
            float(s.total_price or 0),
            float(s.total_received or 0),
            float(s.current_balance or 0),
            s.dealer.name if s.dealer else '-',
            s.commission_status
        ]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
    
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 15

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_dealer_commission_report_excel(sales, dealer=None, start_date=None, end_date=None):
    """Premium Dealer Commission Ledger Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dealer Commission"
    
    MAROON = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    WHITE_BOLD = Font(color="FFFFFF", bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    title = f"DEALER COMMISSION LEDGER"
    if dealer: title += f" - {dealer.name.upper()}"
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color="7D0541")
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Date', 'Dealer', 'Project', 'Plot No', 'Plot Price', 'Comm %', 'Comm Value', 'Paid', 'Remaining', 'Status']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = WHITE_BOLD; c.fill = MAROON; c.border = BORDER; c.alignment = Alignment(horizontal='center')

    for r_idx, s in enumerate(sales, 4):
        remaining = (s.dealer_commission or 0) - (s.dealer_paid_amount or 0)
        data = [
            s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '-',
            s.dealer.name if s.dealer else '-',
            s.plot.project.name if s.plot and s.plot.project else '-',
            s.plot.plot_number if s.plot else '-',
            float(s.total_price or 0),
            float(s.dealer.commission_percentage if s.dealer else 0),
            float(s.dealer_commission or 0),
            float(s.dealer_paid_amount or 0),
            float(remaining),
            s.commission_status
        ]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 15

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_client_payment_report_excel(sales=None, start_date=None, end_date=None, ledger_entries=None, summary=None):
    """Premium Client Payment History Excel - Can be summary or detailed ledger"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Ledger" if ledger_entries else "Client Payments"
    
    MAROON = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    WHITE_BOLD = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if ledger_entries and summary:
        # DETAILED LEDGER MODE
        ws.merge_cells('A1:F1')
        ws['A1'] = "CUSTOMER TRANSACTION LEDGER"
        ws['A1'].font = Font(size=14, bold=True, color="7D0541")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.cell(row=3, column=1, value="Customer:").font = BOLD_FONT
        ws.cell(row=3, column=2, value=summary['customer_name'])
        ws.cell(row=3, column=4, value="Phone:").font = BOLD_FONT
        ws.cell(row=3, column=5, value=summary['customer_phone'])
        
        ws.cell(row=4, column=1, value="Outstanding Balance:").font = BOLD_FONT
        ws.cell(row=4, column=2, value=summary['outstanding_balance']).number_format = '#,##0.00'
        
        headers = ['Date', 'Description', 'Type', 'Debit', 'Credit', 'Balance']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=6, column=i, value=h)
            c.font = WHITE_BOLD; c.fill = MAROON; c.border = BORDER; c.alignment = Alignment(horizontal='center')

        for r_idx, e in enumerate(ledger_entries, 7):
            data = [e['date'], e['description'], e['type'], e['debit'], e['credit'], e['balance']]
            for c_idx, val in enumerate(data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = BORDER
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
    else:
        # SUMMARY MODE (Legacy)
        ws.merge_cells('A1:I1')
        ws['A1'] = "CLIENT PAYMENT & DUE REPORT"
        ws['A1'].font = Font(size=14, bold=True, color="7D0541")
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Date', 'Client', 'Phone', 'Project', 'Plot No', 'Total Price', 'Received', 'Balance', 'Status']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = WHITE_BOLD; c.fill = MAROON; c.border = BORDER; c.alignment = Alignment(horizontal='center')

        for r_idx, s in enumerate(sales or [], 4):
            data = [
                s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '-',
                s.customer.name if s.customer else '-',
                s.customer.phone if s.customer else '-',
                s.plot.project.name if s.plot and s.plot.project else '-',
                s.plot.plot_number if s.plot else '-',
                float(s.total_price or 0) if s.total_price is not None else 0.0,
                float(s.total_received or 0) if s.total_received is not None else 0.0,
                float(s.current_balance or 0) if s.current_balance is not None else 0.0,
                s.commission_status
            ]
            for c_idx, val in enumerate(data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = BORDER
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')

    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_plots_report_excel(plots):
    """Premium Plots Inventory Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plots Inventory"
    
    MAROON = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    WHITE_BOLD = Font(color="FFFFFF", bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:H1')
    ws['A1'] = "PLOTS INVENTORY REPORT"
    ws['A1'].font = Font(size=14, bold=True, color="7D0541")
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['S/No', 'Project', 'Plot No', 'Size', 'Price', 'Status', 'Customer', 'Dealer']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = WHITE_BOLD; c.fill = MAROON; c.border = BORDER; c.alignment = Alignment(horizontal='center')

    for r_idx, p in enumerate(plots, 4):
        data = [
            r_idx - 3,
            p.project.name if p.project else '-',
            p.plot_number,
            p.plot_size,
            float(p.total_price or 0),
            p.status,
            p.customer.name if p.customer else '-',
            p.dealer.name if p.dealer else '-'
        ]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 15

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_landowner_payout_report_excel(sales, project=None, start_date=None, end_date=None):
    """Premium Landowner Payout Excel Report"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from decimal import Decimal
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Landowner Payout"
    
    MAROON = PatternFill(start_color="7D0541", end_color="7D0541", fill_type="solid")
    GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    WHITE_BOLD = Font(color="FFFFFF", bold=True)
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:H1')
    ws['A1'] = "LANDOWNER (MALIK) PAYOUT REPORT"
    ws['A1'].font = Font(size=14, bold=True, color="7D0541")
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Date', 'Client', 'Plot No', 'Size', 'Sale Value', 'Commission %', 'Comm Value', 'Received']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = WHITE_BOLD; c.fill = MAROON; c.border = BORDER; c.alignment = Alignment(horizontal='center')

    total_comm = Decimal('0')
    total_received = Decimal('0')

    for r_idx, s in enumerate(sales, 4):
        comm_val = s.landowner_commission or Decimal('0')
        recv_val = s.landowner_commission_received or Decimal('0')
        total_comm += comm_val
        total_received += recv_val
        
        data = [
            s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '-',
            s.customer.name if s.customer else '-',
            s.plot.plot_number if s.plot else '-',
            s.plot.plot_size if s.plot else '-',
            float(s.total_price or 0),
            float(getattr(s.plot.project, 'landowner_commission_percentage', 0) if s.plot and s.plot.project else 0),
            float(comm_val),
            float(recv_val)
        ]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    # Add Totals
    curr = 4 + len(sales)
    ws.cell(row=curr, column=6, value="TOTAL").font = Font(bold=True)
    ws.cell(row=curr, column=7, value=float(total_comm)).font = Font(bold=True)
    ws.cell(row=curr, column=8, value=float(total_received)).font = Font(bold=True)
    for c in range(6, 9):
        ws.cell(row=curr, column=c).border = BORDER
        ws.cell(row=curr, column=c).fill = GOLD

    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 15

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

